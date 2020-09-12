import requests
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import datetime
from email.mime.text import MIMEText
import smtplib
import random
import sys
import getopt


# 从xlsx文件中读取数据，未使用
def get_info_from_xlsx():
    username = []
    password = []
    userEmail = []
    wb = load_workbook(filename='config.xlsx')
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    row = wb[sheetnames[0]].max_row
    for rowNum in range(1, row + 1):
        # print(sheet.cell(row=rowNum, column=1).value)
        username.append(sheet.cell(row=rowNum, column=1).value)
        # print(sheet.cell(row=rowNum, column=2).value)
        password.append(sheet.cell(row=rowNum, column=2).value)
        # print(sheet.cell(row=rowNum, column=3).value)
        userEmail.append(sheet.cell(row=rowNum, column=3).value)
    sum = len(username)
    return username, password, userEmail, sum


# 从txt文件中读取数据
def get_info_from_txt():
    username = []
    password = []
    userEmail = []
    name = []
    sum = 0
    configFile = open('config.txt', 'r', encoding='utf-8')
    list = configFile.readlines()
    # print(list)
    for lists in list:
        lists = lists.strip()
        lists = lists.strip('\n')
        lists = lists.strip(',')
        lists = lists.split(',')
        # print(lists[0])
        username.append(lists[0])
        # print(lists[1])
        password.append(lists[1])
        # print(lists[2])
        userEmail.append(lists[2])
        # print(lists[3])
        name.append(lists[3])
        sum += 1
    # print('sum: ', sum)
    return username, password, userEmail, name, sum


# 登录，为后续做准备
def login(s, headers, username, password):
    login_url = 'http://yiqing.ctgu.edu.cn/wx/index/loginSubmit.do'
    data = {
        'username': '',
        'password': '',
    }
    data['username'] = username
    data['password'] = password
    r = s.post(url=login_url, headers=headers, data=data)
    # 增加等待时间，让数据接收完毕
    time.sleep(20)
    return r.text


# 获取要带有data的html
def get_student_info(s, headers):
    student_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/toApply.do'
    r = s.get(url=student_info_url, headers=headers)
    # print(r.text)
    # 增加等待时间，让数据接收完毕
    time.sleep(20)
    return r.text


# 获取上报成功的html
def get_success_send_info(s, headers):
    success_send_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/main.do'
    r = s.get(url=success_send_info_url, headers=headers)
    # print(r.text)
    time.sleep(30)
    return r.text


# 解析html,获得data数据
def student_info_parse(html):
    bs = BeautifulSoup(html, 'lxml')
    data = {
        'ttoken': bs.find(attrs={'name': 'ttoken'})['value'],
        # 省份
        'province': bs.find(attrs={'name': 'province'})['value'],
        # 城市
        'city': bs.find(attrs={'name': 'city'})['value'],
        # 县区
        'district': bs.find(attrs={'name': 'district'})['value'],
        # 地区代码：身份证号前六位
        'adcode': bs.find(attrs={'name': 'adcode'})['value'],
        'longitude': bs.find(attrs={'name': 'longitude'})['value'],
        'latitude': bs.find(attrs={'name': 'latitude'})['value'],
        # 是否确诊新型肺炎
        'sfqz': bs.find(attrs={'name': 'sfqz'})['value'],
        # 是否疑似感染
        'sfys': bs.find(attrs={'name': 'sfys'})['value'],
        'sfzy': bs.find(attrs={'name': 'sfzy'})['value'],
        # 是否隔离
        'sfgl': bs.find(attrs={'name': 'sfgl'})['value'],
        # 状态，statusName: "正常"
        'status': bs.find(attrs={'name': 'status'})['value'],
        'sfgr': bs.find(attrs={'name': 'sfgr'})['value'],
        'szdz': bs.find(attrs={'name': 'szdz'})['value'],
        # 手机号
        'sjh': bs.find(attrs={'name': 'sjh'})['value'],
        # 紧急联系人姓名
        'lxrxm': bs.find(attrs={'name': 'lxrxm'})['value'],
        # 紧急联系人手机号
        'lxrsjh': bs.find(attrs={'name': 'lxrsjh'})['value'],
        # 是否发热
        'sffr': '否',
        'sffy': '否',
        'sfgr': '否',
        'qzglsj': '',
        'qzgldd': '',
        'glyy': '',
        'mqzz': '',
        # 是否返校
        'sffx': '否',
        # 其它
        'qt': '',
    }
    # print(data)
    return data


# 解析html，获取success_send_info数据
def success_send_info_parse(html):
    bs = BeautifulSoup(html, 'lxml')
    text = bs.find('span', {"class": "normal-sm-tip green-warn fn-ml10"}).get_text()
    # print(text)
    return text


# 向服务器post数据
def sent_info(s, headers, data):
    sent_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/saveApply.do'
    r = s.post(url=sent_info_url, headers=headers, data=data)
    # print(r.text)
    # 增加等待时间，让数据接收完毕
    time.sleep(30)
    print(r.status_code)


# 发邮件通知结果
def send_rusult(text, fromEmail, passworld, userEmail):
    if userEmail == 'null':
        print('用户指定不发送邮件')
        return
    # nowtime = datetime.datetime.now().strftime('%m-%d')
    # msg_from = ''  # 发送邮箱
    msg_from = fromEmail
    # passwd = ''  # 密码
    passwd = passworld
    msg_to = userEmail  # 目的邮箱

    subject = '安全上报结果'
    # content = text + '\n' + nowtime
    content = text
    msg = MIMEText(content)
    msg['Subject'] = subject
    msg['From'] = msg_from
    msg['To'] = msg_to

    s = smtplib.SMTP_SSL("smtp.qq.com", 465)
    s.login(msg_from, passwd)
    s.sendmail(msg_from, msg_to, msg.as_string())
    print("邮件发送成功")
    s.quit()


def parse_options(argv):
    fromEmail = ''
    pop3Key = ''
    helpMsg = "send.py\n" \
              "\tsend.py -f <fromEmail> -k <pop3Key>\n" \
              "\t-f\t\t\t\t邮件发送方邮箱地址\n" \
              "\t--fromEmail\n" \
              "\t-k\t\t\t\t邮箱POP3授权码\n" \
              "\t--pop3Key\n" \
              "\t-h\t\t\t\thelp\n" \
              "\t--help\n"

    if not ((('-f' in argv) or ('--fromEmail' in argv)) and (('-k' in argv) or ('--pop3Key' in argv))):
        print(helpMsg)
        sys.exit()

    argc = 0
    if ('-h' in argv) or ('--help' in argv):
        argc += 1
    if ('-f' in argv) or ('--fromEmail' in argv):
        argc += 2
    if ('-f' in argv) or ('--fromEmail' in argv):
        argc += 2
    if len(argv) != argc:
        print(helpMsg)
        sys.exit()

    try:
        opts, args = getopt.getopt(argv, "hf:k:", ["help", "fromEmail=", "pop3Key="])
    except getopt.GetoptError:
        print(helpMsg)
        sys.exit(2)

    if not len(opts):
        print(helpMsg)
        sys.exit()

    for option, value in opts:
        if option in ("-h", "--help"):
            print(helpMsg)
        elif option in ("-f", "--fromEmail"):
            if value is None:
                sys.exit()
            fromEmail = value
        elif option in ("-k", "--pop3Key"):
            if value is None:
                sys.exit()
            pop3Key = value
    return fromEmail, pop3Key


def main(argv):
    fromEmail, pop3Key = parse_options(argv)
    username, password, userEmail, name, sum = get_info_from_txt()
    # username, password, userEmail, sum = get_info_from_xlsx()
    print('---------------------------')
    print('账号密码读取成功，共', sum, '人')
    print('---------------------------\n')
    finished = 0
    reported = 0
    # 发送给用户的邮件信息
    userString = []
    # 发送给admin的邮件信息
    adminString = []

    adminString.append('---------------------------')
    adminString.append('账号密码读取成功，共' + str(sum) + '人')
    adminString.append('---------------------------' + '\n')
    try:
        for i in range(sum):
            print('')
            headers = {
                'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Mobile Safari/537.36',
                'Connection': 'close',
            }
            s = requests.session()
            s.keep_alive = False
            # print(s)
            # print(headers)
            state = login(s, headers, username[i], password[i])
            if state == 'success':
                print('用户', name[i], username[i], '登录成功')
                # 记录用户登录成功，添加到userString和adminString
                text = '用户 ' + str(name[i]) + str(username[i]) + '登录成功'
                userString.append(text)
                adminString.append(text)
            else:
                print('用户', name[i], username[i], '密码错误，登录失败')
                # 记录用户登录失败，添加到userString和adminString
                text = '用户 ' + str(name[i]) + str(username[i]) + '密码错误，登录失败'
                userString.append(text)
                adminString.append(text)
                # 将登录失败信息 邮件发送给用户
                send_rusult('\n'.join(userString), fromEmail, pop3Key, userEmail[i])
                # 调用clear()方法，清空列表，避免其出现在下一个用户的邮件中
                # print('userString:---------  : ', userString)
                userString.clear()
                continue
            html = get_student_info(s, headers)
            try:
                data = student_info_parse(html)
                sent_info(s, headers, data)
                # 从网页中解析是否上报成功
                html = get_success_send_info(s, headers)
                msg = success_send_info_parse(html)
                if msg == '今日已上报':
                    # 记录用户上报成功，添加到userString和adminString
                    text = '用户 ' + name[i] + '-' + str(username[i]) + ' 上报成功'
                    print(text)
                    userString.append(text)
                    adminString.append(text + '\n')
                    # print(userEmail[i])
                    # send_rusult(text, userEmail[i])
                    # 将用户上报成功信息，邮件发送给用户
                    send_rusult('\n'.join(userString), fromEmail, pop3Key, userEmail[i])
                    # 调用clear()方法，清空列表，避免其出现在下一个用户的邮件中
                    # print('userString:---------  : ', userString)
                    userString.clear()
                    # 本次上报人数和今日已上报人数统计
                    reported += 1
                    finished += 1
            except:
                # 从网页中解析是否上报成功
                html = get_success_send_info(s, headers)
                msg = success_send_info_parse(html)
                if msg == '今日已上报':
                    text = '用户 ' + name[i] + '-' + str(username[i]) + ' 今日已上报'
                    print(text)
                    adminString.append(text + '\n')
                    # print(userEmail[i])
                    # send_rusult(text, userEmail[i])
                    # 今日已上报统计
                    reported += 1
            # 调用clear()方法，清空列表，避免其出现在下一个用户的邮件中
            # print('userString:---------  : ', userString)
            userString.clear()
            # 增大等待间隔，github访问速度慢
            time.sleep(random.randint(20, 50))
            # time.sleep(random.randint(1, 10))
    finally:
        text = '应报:' + str(sum) + ' 本次上报:' + str(finished) + ' 今日已上报:' + str(reported)
        adminString.append('---------------------------')
        adminString.append(text)
        adminString.append('---------------------------')
        # print('\n'.join(adminString))
        print('---------------------------')
        print(text)
        # send_rusult(text, '管理员邮箱地址')
        send_rusult('\n'.join(adminString), fromEmail, pop3Key, fromEmail)
        print('---------------------------\n')
        # 调用clear()方法，清空列表，避免其出现在下次的邮件中
        # print('adminString:---------  : ', adminString)
        userString.clear()


if __name__ == "__main__":
    main(sys.argv[1:])
